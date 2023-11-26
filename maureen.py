import openpyxl

def readtxt(): 
    with open('saf_field.txt' , 'r' , encoding='utf-8') as f:
        contents = f.readlines()
    print(contents[2],contents[0],contents[1])
    dam = str(89254021324219626518)

    return dam


def searchxlsx(jay):
    print("Number to be found " + jay)
    njia = 'saf_excel.xlsx'
    wb_obj = openpyxl.load_workbook(njia)
    sheet_obj = wb_obj.active 
    column = sheet_obj.max_column 
    row = sheet_obj.max_row

    print("Total Rows:", row) 
    print("Total Columns:", column)

    for i in range(1,row+1):
        cell_obj = sheet_obj.cell(row = i,column = 3)
        if jay == str(cell_obj.value):
            cell_objj = sheet_obj.cell(row=i,column=5)
            print(jay + " found in row "+ str(i))
            print("Top up amount is :"+ str(cell_objj.value))
        else:
            pass
            
ora = readtxt()
searchxlsx(ora)
